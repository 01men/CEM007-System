"""Excel 导入导出引擎（PRD 第 9 章 + FR-4/FR-9.1）。

三类模板由服务端按当前字典动态生成：
- energy：能耗台账导入模板
- carbon：碳活动数据导入模板
- factor：排放因子导入模板（仅管理员）

通用规范：第 1 行表头（冻结），第 2 行灰色示例行（解析时忽略），第 3 行起数据；
必填列表头红色字体；Sheet2 为"填写说明与字典"。
"""
import json
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from . import calc, config, db

RED = Font(color="FF0000", bold=True)
GREY = Font(color="999999", italic=True)
HEADER_FILL = PatternFill("solid", fgColor="F0F2F7")

BIZ_TYPES = ("energy", "carbon", "factor")


# ── 模板定义：列 = (表头, 是否必填, 示例值) ──

def _template_columns(biz_type: str) -> list[tuple[str, bool, object]]:
    if biz_type == "energy":
        return [("年度", True, 2025), ("月份", True, 1), ("能源类型编码", True, "ELEC"),
                ("用能单元编码", True, "PLANT"), ("用量", True, 295000),
                ("费用（元）", False, 210000), ("备注", False, "示例行，导入时自动忽略")]
    if biz_type == "carbon":
        return [("盘查年度", True, 2025), ("科目编码", True, "S2-ELEC-001"),
                ("科目名称", False, "全厂用电量（净）"), ("活动数据", True, 3500000),
                ("备注", False, "示例行，导入时自动忽略")]
    # factor
    return [("科目编码", True, "S2-ELEC-001"), ("因子值", True, 0.0007035),
            ("生效年度起", True, 2025), ("生效年度止", False, ""),
            ("来源出处", True, "华东区域电网平均排放因子 0.7035 tCO₂e/MWh"),
            ("变更原因", True, "年度因子更新")]


def build_template(biz_type: str) -> bytes:
    """生成导入模板（字典与系统实时一致，FR-4.1）。"""
    if biz_type not in BIZ_TYPES:
        raise ValueError(f"未知模板类型：{biz_type}")
    wb = Workbook()
    ws = wb.active
    ws.title = "数据填写"
    cols = _template_columns(biz_type)
    for ci, (head, required, example) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=head)
        cell.font = RED if required else Font(bold=True)
        cell.fill = HEADER_FILL
        ex = ws.cell(row=2, column=ci, value=example)
        ex.font = GREY
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(head)) * 2 + 6, 14)
    ws.freeze_panes = "A2"

    # Sheet2：填写说明与字典
    ws2 = wb.create_sheet("填写说明与字典")
    ws2.append(["填写说明"])
    notes = [
        "1. 第 2 行为示例行（灰色），导入时自动忽略；请从第 3 行起填写数据。",
        "2. 红色表头为必填列；数值列禁止千分位符号；年月为整数。",
        "3. 编码类字段必须严格使用下方字典中的编码，大小写敏感。",
        "4. 同一文件重复导入时，「跳过」策略不会产生重复数据。",
    ]
    if biz_type == "carbon":
        notes.append("5. 科目名称列仅供阅读，导入以科目编码为准。")
    if biz_type == "factor":
        notes.append("5. 生效年度区间与已有版本不得重叠；变更原因为审计必填。")
    for n in notes:
        ws2.append([n])
    ws2.append([])
    if biz_type == "energy":
        ws2.append(["能源类型字典"])
        ws2.append(["编码", "名称", "单位"])
        for r in db.query("SELECT code,name,unit FROM energy_type WHERE enabled=1 ORDER BY code"):
            ws2.append([r["code"], r["name"], r["unit"]])
        ws2.append([])
        ws2.append(["用能单元字典"])
        ws2.append(["编码", "名称"])
        for r in db.query("SELECT code,name FROM energy_unit WHERE enabled=1 ORDER BY id"):
            ws2.append([r["code"], r["name"]])
    else:
        ws2.append(["核算科目字典"])
        ws2.append(["编码", "名称", "范围", "单位", "填报指南摘要"])
        for r in db.query("SELECT code,name_zh,scope,unit,guide FROM emission_source "
                          "WHERE enabled=1 ORDER BY sort_no"):
            ws2.append([r["code"], r["name_zh"], r["scope"], r["unit"],
                        (r["guide"] or "")[:80]])
    for col in "ABCDE":
        ws2.column_dimensions[col].width = 28
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 解析与校验 ──

def _rows_of(file_bytes: bytes) -> list[list]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):  # 跳过表头与示例行
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        rows.append(list(r))
    return rows


def _num(v) -> float | None:
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _int(v) -> int | None:
    n = _num(v)
    return int(n) if n is not None and n == int(n) else None


def parse_and_validate(biz_type: str, file_bytes: bytes) -> dict:
    """全量校验，错误一次性汇总（FR-4.2）。

    返回 {rows: [合法行...], errors: [{row_no, reason, raw}...]}
    row_no 为 Excel 实际行号（数据从第 3 行起）。
    """
    raw_rows = _rows_of(file_bytes)
    valid, errors = [], []
    for i, raw in enumerate(raw_rows):
        row_no = i + 3
        try:
            if biz_type == "energy":
                valid.append(_validate_energy_row(raw))
            elif biz_type == "carbon":
                valid.append(_validate_carbon_row(raw))
            else:
                valid.append(_validate_factor_row(raw))
        except ValueError as exc:
            errors.append({"row_no": row_no, "reason": str(exc), "raw": raw})
    return {"rows": valid, "errors": errors}


def _validate_energy_row(raw: list) -> dict:
    year, month = _int(raw[0] if len(raw) > 0 else None), _int(raw[1] if len(raw) > 1 else None)
    if year is None or not 2000 <= year <= 2100:
        raise ValueError("年度非法（须为 2000-2100 的整数）")
    if month is None or not 1 <= month <= 12:
        raise ValueError("月份非法（须为 1-12 的整数）")
    etype = str(raw[2] or "").strip() if len(raw) > 2 else ""
    if not db.query_one("SELECT 1 FROM energy_type WHERE code=? AND enabled=1", (etype,)):
        raise ValueError(f"能源类型编码 {etype!r} 不在字典中")
    ucode = str(raw[3] or "").strip() if len(raw) > 3 else ""
    unit = db.query_one("SELECT id FROM energy_unit WHERE code=? AND enabled=1", (ucode,))
    if not unit:
        raise ValueError(f"用能单元编码 {ucode!r} 不在字典中")
    qty = _num(raw[4] if len(raw) > 4 else None)
    if qty is None or qty < 0:
        raise ValueError("用量非法（须为 ≥0 的数值）")
    amount = _num(raw[5]) if len(raw) > 5 and raw[5] not in (None, "") else None
    if amount is not None and amount < 0:
        raise ValueError("费用不能为负数")
    remark = str(raw[6])[:200] if len(raw) > 6 and raw[6] else None
    return {"year": year, "month": month, "energy_type_code": etype,
            "unit_id": unit["id"], "quantity": qty, "amount": amount, "remark": remark}


def _validate_carbon_row(raw: list) -> dict:
    year = _int(raw[0] if len(raw) > 0 else None)
    if year is None or not 2000 <= year <= 2100:
        raise ValueError("盘查年度非法")
    code = str(raw[1] or "").strip() if len(raw) > 1 else ""
    if not db.query_one("SELECT 1 FROM emission_source WHERE code=? AND enabled=1", (code,)):
        raise ValueError(f"科目编码 {code!r} 不存在或已停用")
    value = _num(raw[3] if len(raw) > 3 else None)
    if value is None or value < 0:
        raise ValueError("活动数据非法（须为 ≥0 的数值）")
    remark = str(raw[4])[:200] if len(raw) > 4 and raw[4] else None
    return {"year": year, "source_code": code, "activity_value": value, "remark": remark}


def _validate_factor_row(raw: list) -> dict:
    code = str(raw[0] or "").strip() if len(raw) > 0 else ""
    if not db.query_one("SELECT 1 FROM emission_source WHERE code=?", (code,)):
        raise ValueError(f"科目编码 {code!r} 不存在")
    factor = _num(raw[1] if len(raw) > 1 else None)
    if factor is None or factor <= 0:
        raise ValueError("因子值非法（须为 >0 的数值）")
    year_from = _int(raw[2] if len(raw) > 2 else None)
    if year_from is None:
        raise ValueError("生效年度起非法")
    year_to = _int(raw[3]) if len(raw) > 3 and raw[3] not in (None, "") else None
    if year_to is not None and year_to < year_from:
        raise ValueError("生效年度止不能小于起")
    ref = str(raw[4] or "").strip() if len(raw) > 4 else ""
    if not ref:
        raise ValueError("来源出处必填")
    reason = str(raw[5] or "").strip() if len(raw) > 5 else ""
    if not reason:
        raise ValueError("变更原因必填（审计要求）")
    return {"source_code": code, "factor": factor, "year_from": year_from,
            "year_to": year_to, "ref_source": ref, "change_reason": reason}


def build_error_file(biz_type: str, errors: list[dict]) -> bytes:
    """错误明细标注文件：原始行 + 末列错误原因（FR-4.2）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "错误明细"
    heads = [h for h, _r, _e in _template_columns(biz_type)] + ["错误原因"]
    for ci, h in enumerate(heads, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(ci)].width = 18
    for ri, err in enumerate(errors, start=2):
        for ci, v in enumerate(err["raw"], start=1):
            ws.cell(row=ri, column=ci, value=v)
        ws.cell(row=ri, column=len(heads), value=err["reason"]).font = RED
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 入库 ──

def apply_import(biz_type: str, rows: list[dict], strategy: str, user) -> dict:
    """确认入库：skip（幂等跳过重复）/ overwrite（覆盖）。返回新增/覆盖/跳过计数。"""
    added = overwritten = skipped = 0
    now = db.now_iso()
    with db.transaction() as conn:
        for r in rows:
            if biz_type == "energy":
                dup = conn.execute(
                    "SELECT id, status FROM energy_record WHERE year=? AND month=?"
                    " AND energy_type_code=? AND unit_id=?",
                    (r["year"], r["month"], r["energy_type_code"], r["unit_id"])).fetchone()
                if dup:
                    if strategy == "skip" or dup["status"] in ("submitted", "approved"):
                        skipped += 1
                        continue
                    conn.execute(
                        "UPDATE energy_record SET quantity=?,amount=?,remark=?,status='draft',"
                        "updated_at=?,updated_by=? WHERE id=?",
                        (r["quantity"], r["amount"], r["remark"], now, user["name"], dup["id"]))
                    overwritten += 1
                else:
                    conn.execute(
                        "INSERT INTO energy_record(year,month,energy_type_code,unit_id,quantity,"
                        "amount,remark,status,created_at,created_by,updated_at,updated_by)"
                        " VALUES(?,?,?,?,?,?,?,'draft',?,?,?,?)",
                        (r["year"], r["month"], r["energy_type_code"], r["unit_id"],
                         r["quantity"], r["amount"], r["remark"], now, str(user["id"]),
                         now, user["name"]))
                    added += 1
            elif biz_type == "carbon":
                conn.execute(
                    "INSERT INTO carbon_inventory(year,status,created_at,created_by)"
                    " VALUES(?,'draft',?,?) ON CONFLICT(year) DO NOTHING",
                    (r["year"], now, user["name"]))
                inv = conn.execute("SELECT status FROM carbon_inventory WHERE year=?",
                                   (r["year"],)).fetchone()
                if inv["status"] == "approved":
                    skipped += 1  # 已审定年度锁定，跳过
                    continue
                dup = conn.execute(
                    "SELECT id FROM carbon_activity WHERE year=? AND source_code=?",
                    (r["year"], r["source_code"])).fetchone()
                if dup and strategy == "skip":
                    skipped += 1
                    continue
                conn.execute(
                    "INSERT INTO carbon_activity(year,source_code,activity_value,data_origin,"
                    "remark,reported_by,reported_at,created_at,created_by)"
                    " VALUES(?,?,?,'import',?,?,?,?,?) "
                    "ON CONFLICT(year,source_code) DO UPDATE SET activity_value=excluded.activity_value,"
                    "data_origin='import', remark=excluded.remark, reported_by=excluded.reported_by,"
                    "reported_at=excluded.reported_at, updated_at=excluded.updated_at",
                    (r["year"], r["source_code"], r["activity_value"], r["remark"],
                     user["id"], now, now, user["name"]))
                if dup:
                    overwritten += 1
                else:
                    added += 1
            else:  # factor
                from .routers.carbon import _factor_ranges_overlap
                if _factor_ranges_overlap(r["source_code"], r["year_from"], r["year_to"]):
                    skipped += 1
                    continue
                conn.execute(
                    "INSERT INTO emission_factor(source_code,factor,year_from,year_to,"
                    "ref_source,change_reason,created_at,created_by) VALUES(?,?,?,?,?,?,?,?)",
                    (r["source_code"], r["factor"], r["year_from"], r["year_to"],
                     r["ref_source"], r["change_reason"], now, user["name"]))
                added += 1
    return {"added": added, "overwritten": overwritten, "skipped": skipped}


# ── Excel 活动总账导出（FR-9.1，对齐原型三 Sheet）──

def build_ledger_excel(year: int) -> bytes:
    result = calc.compute_year(year)
    wb = Workbook()

    # Sheet1：摘要
    ws = wb.active
    ws.title = "碳指标摘要"
    base = calc.compute_year(config.BASE_YEAR)
    prev_note = "" if base["total"] else "（无基准数据）"
    ws.append(["金华聚杰电器有限公司 温室气体碳盘查及碳生产率总账"])
    ws.append([f"盘查年度: {year}", "", f"生成时间: {db.now_iso()}"])
    ws.append([])
    ws.append(["指标", f"{year} 年度", f"{config.BASE_YEAR} 基准", "同比变动"])
    delta = f"{(result['total']-base['total'])/base['total']*100:.1f}%" if base["total"] else "—"
    rows = [
        ("碳排总量 (tCO₂e)", round(result["total"], 2), round(base["total"], 2), delta),
        ("年度总产值 (万美金)", result["total_revenue"], base["total_revenue"], "—"),
        ("产值碳强度 (tCO₂e/万美金)", round(result["intensity"], 4),
         round(base["intensity"], 4), "—"),
    ]
    for r in rows:
        ws.append(list(r) + [prev_note] if False else list(r))
    ws.append([])
    ws.append(["范围", "排放量 (tCO₂e)", "占比 (%)"])
    for scope, val in result["groups"].items():
        ws.append([scope, round(val, 3),
                   round(val / result["total"] * 100, 1) if result["total"] else 0])
    ws.append(["合计", round(result["total"], 3), 100.0])
    ws.append(["绿电减碳量 (tCO₂e)", result["green"]["pv_reduction"],
               f"光伏自用电量 {result['green']['pv_quantity']} kWh"])

    # Sheet2：活动明细
    ws2 = wb.create_sheet("细分活动明细账")
    ws2.append(["温室气体排放源", "GHG分类", "源类型", "责任填报部门", "活动层级原始数值",
                "数据计量单位", "排放因子", "核算碳排总量 (tCO₂e)", "数据来源方式",
                "活动数据收集审计指引", "排放因子标准/引用来源"])
    for d in result["details"]:
        ws2.append([d["name_zh"], d["scope"], d["source_type"], d["dept_code"],
                    d["activity_value"], d["unit"], d["factor"], d["emission"],
                    d["data_origin"], d["guide"], d["factor_ref"]])
    for col, w in zip("ABCDEFGHIJK", (24, 10, 10, 12, 16, 12, 12, 18, 10, 60, 50)):
        ws2.column_dimensions[col].width = w

    # Sheet3：多年度汇总 + 因子说明
    ws3 = wb.create_sheet("历年汇总")
    ws3.append(["年度", "范围一", "范围二", "范围三", "总量 (tCO₂e)", "强度", "盘查状态"])
    for s in calc.year_summary():
        ws3.append([s["year"], round(s["groups"]["范围一"], 2), round(s["groups"]["范围二"], 2),
                    round(s["groups"]["范围三"], 2), round(s["total"], 2),
                    round(s["intensity"], 4), s["status"]])
    if result["warnings"]:
        ws3.append([])
        ws3.append(["因子回退提示"])
        for w in result["warnings"]:
            ws3.append([w])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 导入暂存（upload→confirm 两步的载体）──

def save_staging(task_id: int, payload: dict) -> Path:
    path = config.FILES_DIR / f"staging_{task_id}.json"
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return path


def load_staging(task_id: int) -> dict:
    path = config.FILES_DIR / f"staging_{task_id}.json"
    return json.loads(path.read_text(encoding="utf-8"))
