# -*- coding: utf-8 -*-
"""一次性构建脚本：解析 data/seed_excel/ 下两份 Excel，生成 server/seed_data.py。

用法：.venv/Scripts/python.exe scripts/build_seed_data.py
数据源：
  - 总表.xlsx（2025年聚杰减碳项目_总表）：活动数据（能耗台账 + 碳活动 + 财务/客户）
  - 排放因子.xlsx（GHG Protocol 排放因子明细）：因子出处标注
总表「计算方式」「项目目录」sheet 为因子口径权威来源。
"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = ROOT / "data" / "seed_excel"
OUT = ROOT / "server" / "seed_data.py"

MONTHS = list(range(1, 13))


def num(v):
    """单元格转 float；空/文本返回 None。"""
    if v is None or isinstance(v, str):
        return None
    return float(v)


def load():
    return openpyxl.load_workbook(EXCEL_DIR / "总表.xlsx", read_only=True, data_only=True)


def parse_natural_gas(ws):
    """天然气 sheet：行=月份(3-14)，列 B-F = 2021/2022/2023/2024/2025，单位 Nm³。"""
    years = {2021: 1, 2022: 2, 2023: 3, 2024: 4, 2025: 5}
    rows = list(ws.iter_rows(min_row=3, max_row=14, values_only=True))
    out = {}  # year -> [12]
    for year, col in years.items():
        vals = [num(r[col]) for r in rows]
        if any(v for v in vals):
            out[year] = [v or 0.0 for v in vals]
    return out


def parse_lpg(ws):
    """液化石油气 sheet：行 5-16，B=油量(升) C=金额(元)，2025 年。"""
    rows = list(ws.iter_rows(min_row=5, max_row=16, values_only=True))
    return [(num(r[1]) or 0.0, num(r[2])) for r in rows]  # [(liters, amount)]


def parse_electricity(ws):
    """用电量统计 sheet：2021(行4-15) 2022(行20-31) 为总表口径；2025(行36-46) 分表，
    G 列「聚杰」为扣减租赁后的自有用电，B 列总度数。1-2月合并数据均摊。"""
    rows = list(ws.iter_rows(values_only=True))
    out = {}  # year -> [(qty, amount, gross, lease)] x 12
    y2021 = [(num(r[1]) or 0.0, num(r[2])) for r in rows[3:15]]
    y2022 = [(num(r[1]) or 0.0, num(r[2])) for r in rows[19:31]]
    out[2021] = [(q, a, q, 0.0) for q, a in y2021]
    out[2022] = [(q, a, q, 0.0) for q, a in y2022]
    y2025 = []
    for r in rows[35:46]:
        label = str(r[0] or "")
        gross = num(r[1]) or 0.0
        own = num(r[6]) or 0.0  # 聚杰列
        if label.startswith("1-2"):
            half_g, half_o = gross / 2, own / 2
            y2025.append((half_o, None, half_g, half_g - half_o))
            y2025.append((half_o, None, half_g, half_g - half_o))
        else:
            y2025.append((own, None, gross, gross - own))
    out[2025] = y2025
    return out


FUEL_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")


def parse_fuel_detail(ws):
    """中国石化柴汽油明细：业务类型=加油 的行，按 月份×品种 归集升数与金额。"""
    agg = {}  # (month, 'DIESEL'|'GASOLINE') -> [liters, amount]
    for r in ws.iter_rows(values_only=True):
        t, biz, kind, qty, _price, amount = (list(r) + [None] * 6)[:6]
        if biz != "加油" or not isinstance(t, str) or not FUEL_RE.match(t):
            continue
        month = int(t[5:7])
        fuel = "DIESEL" if "柴油" in str(kind) else "GASOLINE"
        key = (month, fuel)
        slot = agg.setdefault(key, [0.0, 0.0])
        slot[0] += num(qty) or 0.0
        slot[1] += num(amount) or 0.0
    return agg


def parse_septic_people(ws):
    """化粪池工时 sheet 行4：月平均人数（2021/2022/2024/2025）。"""
    rows = list(ws.iter_rows(values_only=True))
    years = [2021, 2022, 2024, 2025]
    people_row = rows[3]  # 月平均人数
    return {y: num(people_row[i + 1]) for i, y in enumerate(years)}


def parse_private_car_liters(ws):
    """私车公用 sheet「历年费用汇总」：行36-42，A=年份，O=汽油升数。"""
    out = {}
    for r in ws.iter_rows(min_row=36, max_row=42, values_only=True):
        label = str(r[0] or "")
        m = re.match(r"(\d{4})年", label)
        if m:
            out[int(m.group(1))] = num(r[14])  # O 列 汽油升数
    return out


def parse_freight(ws):
    """成品运输-汇总：行2=工厂吨公里，行9=外购吨公里。"""
    rows = list(ws.iter_rows(values_only=True))
    plant = num(rows[1][10]) or 0.0
    outsource = num(rows[8][10]) or 0.0
    return plant, outsource, plant + outsource


def parse_office_refrigerant_add(ws):
    """空调制冷剂（办公）：新增记录表 2025 年 R-32 加注 90kg（行17，GWP 675）。"""
    rows = list(ws.iter_rows(values_only=True))
    for r in rows:
        if str(r[0] or "").strip() == "1" and "R-32" in str(r[1] or ""):
            return num(r[6]) or 0.0, num(r[7]) or 675.0
    return 0.0, 675.0


def main():
    wb = load()
    ng = parse_natural_gas(wb["天然气"])
    lpg_2025 = parse_lpg(wb["液化石油气"])
    elec = parse_electricity(wb["用电量统计"])
    fuel = parse_fuel_detail(wb["中国石化柴汽油明细"])
    people = parse_septic_people(wb["化粪池工时"])
    car_liters = parse_private_car_liters(wb["私车公用"])
    freight_plant, freight_out, freight_total = parse_freight(wb["成品运输-汇总"])
    ref_off_kg, ref_off_gwp = parse_office_refrigerant_add(wb["空调制冷剂（办公）"])

    diesel_2025 = sum(v[0] for (m, f), v in fuel.items() if f == "DIESEL")
    gasoline_2025 = sum(v[0] for (m, f), v in fuel.items() if f == "GASOLINE")
    # 明细未覆盖全部油卡，按「中国石化柴汽油-汇总」权威总量等比缩放月度归集
    SUMMARY_TOTALS = {"DIESEL": 4106.76, "GASOLINE": 19158.29}
    for ftype, current in (("DIESEL", diesel_2025), ("GASOLINE", gasoline_2025)):
        total = SUMMARY_TOTALS[ftype]
        if current and abs(current - total) > 0.01:
            ratio = total / current
            for key in [k for k in fuel if k[1] == ftype]:
                fuel[key][0] *= ratio
                fuel[key][1] *= ratio
    diesel_2025 = SUMMARY_TOTALS["DIESEL"]
    gasoline_2025 = SUMMARY_TOTALS["GASOLINE"]

    # ── 能耗台账（energy_record）：(year, month, type, qty, amount, gross, lease, pv, remark) ──
    records = []
    for year, months in sorted(ng.items()):
        for i, q in enumerate(months, 1):
            if q:
                records.append((year, i, "NG", round(q, 2), None, round(q, 2), 0.0, 0.0,
                                "天然气用量统计表"))
    for i, (q, amt) in enumerate(lpg_2025, 1):
        records.append((2025, i, "LPG", round(q, 2), amt, round(q, 2), 0.0, 0.0,
                        "液化石油气用量统计表"))
    for year, months in sorted(elec.items()):
        for i, (q, amt, gross, lease) in enumerate(months, 1):
            remark = "用电量统计"
            if year == 2025:
                remark = "用电量统计（聚杰分表，总表扣减租赁）"
                if i <= 2:
                    remark += "；1-2月合并均摊"
            records.append((year, i, "ELEC", round(q, 2), amt, round(gross, 2),
                            round(lease, 2), 0.0, remark))
    for (month, ftype), (liters, amount) in sorted(fuel.items()):
        records.append((2025, month, ftype, round(liters, 2), round(amount, 2),
                        round(liters, 2), 0.0, 0.0, "中石化油卡明细归集"))

    # ── 2025 碳活动数据（carbon_activity，manual）──
    activity_2025 = {
        "S1-NG-001": sum(ng.get(2025, [0.0])),          # 2025 新工艺停用 → 0
        "S1-LPG-001": round(sum(q for q, _ in lpg_2025), 2),
        "S1-DIESEL-001": round(diesel_2025, 2),
        "S1-GAS-001": round(gasoline_2025, 2),
        "S1-CO2-001": 0.0,                               # 2025 无 CO2 灭火器补充
        "S1-REF-OFF": ref_off_kg,                        # R-32 补充加注 90kg
        "S1-REF-PRD": 0.0,                               # 2025 无补充加注记录
        "S1-WORKHRS": people[2025],                      # 当量人口 890 人
        "S2-ELEC-001": round(sum(m[0] for m in elec[2025]), 2),  # 聚杰口径净电
        "S3-FREIGHT": round(freight_total, 2),           # 工厂+外购 吨·公里
        "S3-MILEAGE": round(car_liters[2025], 2),        # 私车公用汽油升数
    }

    data = {
        "meta": {
            "source_files": ["总表.xlsx", "排放因子.xlsx"],
            "inventory_year": 2025,
            "total_revenue": 6770.81,  # 万美金（项目目录·实际年产值）
        },
        "energy_records": records,
        "carbon_activity_2025": activity_2025,
        "customers": [
            {"code": "BOSCH", "name_zh": "博世", "name_en": "Bosch", "revenue": 2051.48},
            {"code": "BD", "name_zh": "百得", "name_en": "Black & Decker", "revenue": 265.66},
        ],
        "history": {  # 备查：历史年活动值（未入库，仅供核对）
            "ng": ng, "septic_people": people, "car_liters": car_liters,
            "freight": {"plant": freight_plant, "outsource": freight_out},
        },
    }

    header = '"""预置数据：由 scripts/build_seed_data.py 从 data/seed_excel/ 两份 Excel 解析生成，请勿手改。"""\n'
    OUT.write_text(header + "DATA = " + repr(data) + "\n", encoding="utf-8")

    print(f"能耗台账 {len(records)} 条；2025 柴油 {diesel_2025:.2f}L / 汽油 {gasoline_2025:.2f}L")
    print("2025 碳活动：")
    for k, v in activity_2025.items():
        print(f"  {k}: {v}")
    print(f"已生成 {OUT}")


if __name__ == "__main__":
    main()
